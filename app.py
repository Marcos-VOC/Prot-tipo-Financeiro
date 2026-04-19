# ============================================================
#  Santos & Associados – Sistema de Controle Financeiro
#  app.py  |  Backend principal em Flask + SQLite
# ============================================================

from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from datetime import datetime, date
import sqlite3, os, io

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), "financeiro.db")

# ----------------------------------------------------------
# BANCO DE DADOS
# ----------------------------------------------------------

def get_db():
    """Abre conexão com o banco e retorna linhas como dicionário."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Cria a tabela de movimentações se ainda não existir."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS movimentacoes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo        TEXT    NOT NULL CHECK(tipo IN ('Entrada','Saida')),
                categoria   TEXT    NOT NULL,
                descricao   TEXT    NOT NULL,
                valor       REAL    NOT NULL CHECK(valor > 0),
                data        TEXT    NOT NULL   -- formato YYYY-MM-DD
            )
        """)
        conn.commit()


# ----------------------------------------------------------
# CATEGORIAS (baseadas na planilha do escritório)
# ----------------------------------------------------------

CATEGORIAS = {
    "Entrada": [
        "Honorários – Consultoria",
        "Honorários – Contencioso",
        "Honorários – Trabalhista",
        "Acordos e Êxitos",
        "Pareceres e Assessoria",
        "Outras Receitas",
    ],
    "Saida": [
        "Salários e Encargos",
        "Aluguel e Condomínio",
        "Softwares Jurídicos",
        "Marketing e Publicidade",
        "Custas e Despesas Processuais",
        "Treinamento e Capacitação",
        "Despesas Administrativas",
        "Outras Despesas",
    ],
}

MESES_PT = {
    1:"Janeiro", 2:"Fevereiro", 3:"Março",    4:"Abril",
    5:"Maio",    6:"Junho",    7:"Julho",     8:"Agosto",
    9:"Setembro",10:"Outubro", 11:"Novembro", 12:"Dezembro",
}

# ----------------------------------------------------------
# ROTAS
# ----------------------------------------------------------

@app.route("/")
def dashboard():
    hoje = date.today()
    mes  = int(request.args.get("mes",  hoje.month))
    ano  = int(request.args.get("ano",  hoje.year))

    with get_db() as conn:
        # Totais do mês selecionado
        def total(tipo):
            row = conn.execute(
                "SELECT COALESCE(SUM(valor),0) AS total FROM movimentacoes "
                "WHERE tipo=? AND strftime('%m',data)=? AND strftime('%Y',data)=?",
                (tipo, f"{mes:02d}", str(ano))
            ).fetchone()
            return row["total"]

        entradas = total("Entrada")
        saidas   = total("Saida")
        liquido  = entradas - saidas

        # Últimas 5 movimentações do mês
        recentes = conn.execute(
            "SELECT * FROM movimentacoes "
            "WHERE strftime('%m',data)=? AND strftime('%Y',data)=? "
            "ORDER BY data DESC, id DESC LIMIT 5",
            (f"{mes:02d}", str(ano))
        ).fetchall()

        # Dados para o gráfico (12 meses do ano selecionado)
        grafico = []
        for m in range(1, 13):
            e = conn.execute(
                "SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes "
                "WHERE tipo='Entrada' AND strftime('%m',data)=? AND strftime('%Y',data)=?",
                (f"{m:02d}", str(ano))
            ).fetchone()["t"]
            s = conn.execute(
                "SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes "
                "WHERE tipo='Saida' AND strftime('%m',data)=? AND strftime('%Y',data)=?",
                (f"{m:02d}", str(ano))
            ).fetchone()["t"]
            grafico.append({"mes": MESES_PT[m][:3], "entradas": e, "saidas": s})

        # Top categorias do mês (entradas)
        top_receitas = conn.execute(
            "SELECT categoria, SUM(valor) AS total FROM movimentacoes "
            "WHERE tipo='Entrada' AND strftime('%m',data)=? AND strftime('%Y',data)=? "
            "GROUP BY categoria ORDER BY total DESC",
            (f"{mes:02d}", str(ano))
        ).fetchall()

        # Top categorias do mês (saídas)
        top_despesas = conn.execute(
            "SELECT categoria, SUM(valor) AS total FROM movimentacoes "
            "WHERE tipo='Saida' AND strftime('%m',data)=? AND strftime('%Y',data)=? "
            "GROUP BY categoria ORDER BY total DESC",
            (f"{mes:02d}", str(ano))
        ).fetchall()

        # Anos disponíveis para o filtro
        anos_rows = conn.execute(
            "SELECT DISTINCT strftime('%Y',data) AS ano FROM movimentacoes ORDER BY ano DESC"
        ).fetchall()
        anos = [r["ano"] for r in anos_rows] or [str(hoje.year)]

    return render_template("dashboard.html",
        entradas=entradas, saidas=saidas, liquido=liquido,
        recentes=recentes, grafico=grafico,
        top_receitas=top_receitas, top_despesas=top_despesas,
        mes=mes, ano=ano, anos=anos,
        meses_pt=MESES_PT,
        mes_nome=MESES_PT[mes],
    )


@app.route("/movimentacoes")
def movimentacoes():
    hoje = date.today()
    mes   = request.args.get("mes",  "")
    ano   = int(request.args.get("ano",  hoje.year))
    tipo  = request.args.get("tipo", "")
    busca = request.args.get("busca","").strip()

    query  = "SELECT * FROM movimentacoes WHERE strftime('%Y',data)=?"
    params = [str(ano)]

    if mes:
        query += " AND strftime('%m',data)=?"
        params.append(f"{int(mes):02d}")
    if tipo:
        query += " AND tipo=?"
        params.append(tipo)
    if busca:
        query += " AND (descricao LIKE ? OR categoria LIKE ?)"
        params += [f"%{busca}%", f"%{busca}%"]

    query += " ORDER BY data DESC, id DESC"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()
        totais = conn.execute(
            "SELECT tipo, COALESCE(SUM(valor),0) AS total FROM movimentacoes "
            "WHERE strftime('%Y',data)=? GROUP BY tipo", [str(ano)]
        ).fetchall()
        anos_rows = conn.execute(
            "SELECT DISTINCT strftime('%Y',data) AS ano FROM movimentacoes ORDER BY ano DESC"
        ).fetchall()
        anos = [r["ano"] for r in anos_rows] or [str(hoje.year)]

    totais_dict = {r["tipo"]: r["total"] for r in totais}

    return render_template("movimentacoes.html",
        rows=rows, mes=mes, ano=ano, tipo=tipo, busca=busca,
        totais=totais_dict, anos=anos, meses_pt=MESES_PT,
    )


@app.route("/nova", methods=["GET","POST"])
def nova():
    erro = None
    if request.method == "POST":
        tipo      = request.form.get("tipo","")
        categoria = request.form.get("categoria","")
        descricao = request.form.get("descricao","").strip()
        valor_str = request.form.get("valor","").replace(",",".")
        data_str  = request.form.get("data","")

        # Validações básicas
        try:
            valor = float(valor_str)
            assert valor > 0
        except:
            erro = "Valor inválido. Use números positivos (ex: 1500.00)."

        if not erro and not descricao:
            erro = "Descrição não pode estar vazia."

        if not erro:
            try:
                datetime.strptime(data_str, "%Y-%m-%d")
            except:
                erro = "Data inválida."

        if not erro:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO movimentacoes (tipo,categoria,descricao,valor,data) VALUES (?,?,?,?,?)",
                    (tipo, categoria, descricao, valor, data_str)
                )
                conn.commit()
            return redirect(url_for("dashboard"))

    hoje_str = date.today().strftime("%Y-%m-%d")
    tipo_pre = request.args.get("tipo","Entrada")
    return render_template("nova.html",
        categorias=CATEGORIAS, hoje=hoje_str,
        tipo_pre=tipo_pre, erro=erro,
    )


@app.route("/excluir/<int:mov_id>", methods=["POST"])
def excluir(mov_id):
    with get_db() as conn:
        conn.execute("DELETE FROM movimentacoes WHERE id=?", (mov_id,))
        conn.commit()
    return redirect(request.referrer or url_for("movimentacoes"))


@app.route("/api/grafico/<int:ano>")
def api_grafico(ano):
    """Retorna JSON com dados mensais para o gráfico."""
    with get_db() as conn:
        dados = []
        for m in range(1, 13):
            e = conn.execute(
                "SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes "
                "WHERE tipo='Entrada' AND strftime('%m',data)=? AND strftime('%Y',data)=?",
                (f"{m:02d}", str(ano))
            ).fetchone()["t"]
            s = conn.execute(
                "SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes "
                "WHERE tipo='Saida' AND strftime('%m',data)=? AND strftime('%Y',data)=?",
                (f"{m:02d}", str(ano))
            ).fetchone()["t"]
            dados.append({"mes": MESES_PT[m][:3], "entradas": e, "saidas": s})
    return jsonify(dados)


# ----------------------------------------------------------
# EXPORTAR — PÁGINA DE CONFIGURAÇÃO
# ----------------------------------------------------------

@app.route("/exportar-config", methods=["GET"])
def exportar_config():
    hoje = date.today()
    with get_db() as conn:
        anos_rows = conn.execute(
            "SELECT DISTINCT strftime('%Y',data) AS ano FROM movimentacoes ORDER BY ano DESC"
        ).fetchall()
        anos = [r["ano"] for r in anos_rows] or [str(hoje.year)]
        todas_cats = conn.execute(
            "SELECT DISTINCT categoria FROM movimentacoes ORDER BY categoria"
        ).fetchall()
        categorias_db = [r["categoria"] for r in todas_cats]

    return render_template("exportar.html",
        anos=anos, hoje=hoje,
        meses_pt=MESES_PT,
        categorias=CATEGORIAS,
        categorias_db=categorias_db,
    )


# ----------------------------------------------------------
# EXPORTAR — GERAÇÃO DO ARQUIVO
# ----------------------------------------------------------

def _estilo(wb):
    """Retorna helpers de estilo reutilizáveis."""
    AZUL_ESCURO  = "0F1F3D"
    DOURADO      = "C9A84C"
    VERDE        = "198754"
    VERMELHO     = "DC3545"
    CINZA_CLARO  = "F4F6FB"
    BORDA_COR    = "DEE2E6"

    thin = Side(style="thin", color=BORDA_COR)
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(ws, row, col, texto, bg=AZUL_ESCURO, fg="FFFFFF", bold=True, size=11):
        c = ws.cell(row=row, column=col, value=texto)
        c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
        return c

    def data_cell(ws, row, col, value, fmt=None, bold=False, color="000000", bg=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=bold, color=color, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = borda
        if fmt:
            c.number_format = fmt
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    return header_cell, data_cell, AZUL_ESCURO, DOURADO, VERDE, VERMELHO, CINZA_CLARO, borda


@app.route("/exportar", methods=["POST"])
def exportar():
    hoje     = date.today()
    fmt_brl  = '"R$" #,##0.00'
    fmt_pct  = '0.0"%"'

    # ── Parâmetros do formulário ─────────────────────────────
    tipo_rel   = request.form.get("tipo_relatorio", "mensal")
    mes_ini    = int(request.form.get("mes_ini",  hoje.month))
    ano_ini    = int(request.form.get("ano_ini",  hoje.year))
    mes_fim    = int(request.form.get("mes_fim",  hoje.month))
    ano_fim    = int(request.form.get("ano_fim",  hoje.year))
    filtro_tipo = request.form.get("filtro_tipo", "")        # Entrada/Saida/""
    cats_sel   = request.form.getlist("categorias")          # lista ou vazia = todas
    inc_det    = request.form.get("incluir_detalhes") == "1"
    ordenar    = request.form.get("ordenar", "data")

    # Para relatório mensal o período é só 1 mês
    if tipo_rel == "mensal":
        mes_fim = mes_ini
        ano_fim = ano_ini

    # ── Query base ───────────────────────────────────────────
    def query_movs(m_ini, a_ini, m_fim, a_fim, tipo=None, cats=None, order="data ASC, id ASC"):
        cond = ["data >= ? AND data <= ?"]
        params = [
            f"{a_ini}-{m_ini:02d}-01",
            f"{a_fim}-{m_fim:02d}-31",
        ]
        if tipo:
            cond.append("tipo = ?"); params.append(tipo)
        if cats:
            ph = ",".join("?" * len(cats))
            cond.append(f"categoria IN ({ph})"); params += cats
        col = {"data":"data ASC, id ASC", "valor":"valor DESC", "categoria":"categoria ASC, data ASC"}.get(order, "data ASC")
        sql = f"SELECT * FROM movimentacoes WHERE {' AND '.join(cond)} ORDER BY {col}"
        with get_db() as conn:
            return conn.execute(sql, params).fetchall()

    def soma_periodo(tipo, m_ini, a_ini, m_fim, a_fim, cats=None):
        cond = ["data >= ?","data <= ?","tipo = ?"]
        params = [f"{a_ini}-{m_ini:02d}-01", f"{a_fim}-{m_fim:02d}-31", tipo]
        if cats:
            ph = ",".join("?" * len(cats)); cond.append(f"categoria IN ({ph})"); params += cats
        with get_db() as conn:
            r = conn.execute(f"SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes WHERE {' AND '.join(cond)}", params).fetchone()
        return r["t"]

    # Helpers de estilo
    wb = Workbook()
    header_cell, data_cell, AZ, DOC, VD, VM, CZ, borda = _estilo(wb)

    def titulo_sheet(ws, titulo, subtitulo, ncols=5):
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 36
        ws.merge_cells(f"B2:{get_column_letter(ncols+1)}2")
        c = ws["B2"]
        c.value     = f"⚖  Santos & Associados  |  {titulo}"
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        c.fill      = PatternFill("solid", fgColor=AZ)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20
        ws.merge_cells(f"B3:{get_column_letter(ncols+1)}3")
        c = ws["B3"]
        c.value     = subtitulo
        c.font      = Font(name="Arial", italic=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=DOC)
        c.alignment = Alignment(horizontal="center", vertical="center")

    def rodape(ws, row, ncols=5):
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f"B{row}:{get_column_letter(ncols+1)}{row}")
        c = ws.cell(row=row, column=2,
                    value="Santos & Associados Advocacia  •  Relatório gerado automaticamente  •  Uso interno")
        c.font      = Font(name="Arial", italic=True, color="AAAAAA", size=8)
        c.alignment = Alignment(horizontal="center")

    def bloco_resumo(ws, row, e, s, ncols=2):
        liq = e - s
        cor_l_bg  = "DBEAFE" if liq >= 0 else "FEE2E2"
        cor_l_txt = "1E40AF" if liq >= 0 else VM
        itens = [
            ("▲  Total de Entradas", e, "D1FAE5", VD),
            ("▼  Total de Saídas",   s, "FEE2E2", VM),
            ("=  Lucro Líquido",     liq, cor_l_bg, cor_l_txt),
        ]
        ws.row_dimensions[row].height = 20
        header_cell(ws, row, 2, "INDICADOR", bg=AZ, size=9)
        header_cell(ws, row, 3, "VALOR (R$)", bg=AZ, size=9)
        if ncols > 2:
            header_cell(ws, row, 4, "MARGEM (%)", bg=AZ, size=9)
        row += 1
        for label, val, bg, cor in itens:
            ws.row_dimensions[row].height = 24
            data_cell(ws, row, 2, label, bold=True, color="333333", bg="F8F9FA")
            data_cell(ws, row, 3, val, fmt=fmt_brl, bold=True, color=cor, bg=bg, align="right")
            if ncols > 2:
                marg = (val / e * 100) if e > 0 and label != "▲  Total de Entradas" else (100.0 if e > 0 else 0)
                data_cell(ws, row, 4, marg if label != "▲  Total de Entradas" else 100.0,
                          fmt=fmt_pct, bold=True, color=cor, bg=bg, align="right")
            row += 1
        return row

    def tabela_movimentacoes(ws, movs, start_row, periodo_label):
        ws.row_dimensions[start_row].height = 20
        cols = ["Data","Tipo","Categoria","Descrição","Valor (R$)"]
        for ci, nome in enumerate(cols, 2):
            header_cell(ws, start_row, ci, nome, size=9)
        r = start_row + 1
        for m in movs:
            ws.row_dimensions[r].height = 18
            bg = "F0FFF4" if m["tipo"] == "Entrada" else "FFF5F5"
            cor = VD if m["tipo"] == "Entrada" else VM
            data_cell(ws, r, 2, datetime.strptime(m["data"],"%Y-%m-%d").strftime("%d/%m/%Y"), bg=bg, align="center")
            data_cell(ws, r, 3, m["tipo"], bold=True, color=cor, bg=bg, align="center")
            data_cell(ws, r, 4, m["categoria"], bg=bg)
            data_cell(ws, r, 5, m["descricao"], bg=bg)
            data_cell(ws, r, 6, m["valor"], fmt=fmt_brl, bold=True, color=cor, bg=bg, align="right")
            r += 1
        # Linha de total
        ws.row_dimensions[r].height = 22
        data_cell(ws, r, 2, "TOTAL", bold=True, bg="E2E8F0", align="center")
        ws.merge_cells(f"C{r}:E{r}")
        data_cell(ws, r, 3, f"{len(movs)} registro(s)", bold=True, bg="E2E8F0", align="center")
        tc = ws.cell(row=r, column=6, value=f"=SUM(F{start_row+1}:F{r-1})")
        tc.font = Font(name="Arial", bold=True, color=AZ, size=10)
        tc.number_format = fmt_brl
        tc.fill = PatternFill("solid", fgColor="E2E8F0")
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border = borda
        return r + 1

    # ════════════════════════════════════════════════════════════
    #  TIPO 1 — RELATÓRIO MENSAL
    # ════════════════════════════════════════════════════════════
    if tipo_rel == "mensal":
        mes_nome = MESES_PT[mes_ini]
        movs = query_movs(mes_ini, ano_ini, mes_fim, ano_fim,
                          filtro_tipo or None, cats_sel or None, ordenar)
        e = soma_periodo("Entrada", mes_ini, ano_ini, mes_fim, ano_fim, cats_sel or None)
        s = soma_periodo("Saida",   mes_ini, ano_ini, mes_fim, ano_fim, cats_sel or None)

        ws = wb.active; ws.title = "Resumo Mensal"
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

        titulo_sheet(ws, "Relatório Mensal", f"{mes_nome}/{ano_ini}  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=2)

        r = bloco_resumo(ws, 5, e, s, ncols=3)
        r += 1

        # Por categoria
        with get_db() as conn:
            por_cat = conn.execute(
                "SELECT tipo, categoria, SUM(valor) AS total FROM movimentacoes "
                "WHERE data >= ? AND data <= ? GROUP BY tipo, categoria ORDER BY tipo, total DESC",
                (f"{ano_ini}-{mes_ini:02d}-01", f"{ano_ini}-{mes_ini:02d}-31")
            ).fetchall()

        ws.row_dimensions[r].height = 20
        header_cell(ws, r, 2, "CATEGORIA", bg=AZ, size=9)
        header_cell(ws, r, 3, "TOTAL (R$)", bg=AZ, size=9)
        header_cell(ws, r, 4, "% DO TIPO", bg=AZ, size=9)
        r += 1
        tipo_ult = None
        for pc in por_cat:
            if pc["tipo"] != tipo_ult:
                tipo_ult = pc["tipo"]
                bg_h = VD if tipo_ult == "Entrada" else VM
                ws.row_dimensions[r].height = 18
                header_cell(ws, r, 2, "▲ RECEITAS" if tipo_ult == "Entrada" else "▼ DESPESAS", bg=bg_h, size=9)
                header_cell(ws, r, 3, "", bg=bg_h); header_cell(ws, r, 4, "", bg=bg_h)
                r += 1
            total_tipo = e if pc["tipo"] == "Entrada" else s
            bg = "F0FFF4" if pc["tipo"] == "Entrada" else "FFF5F5"
            cor = VD if pc["tipo"] == "Entrada" else VM
            ws.row_dimensions[r].height = 18
            data_cell(ws, r, 2, pc["categoria"], bg=bg)
            data_cell(ws, r, 3, pc["total"], fmt=fmt_brl, color=cor, bg=bg, align="right")
            pct = (pc["total"]/total_tipo*100) if total_tipo > 0 else 0
            data_cell(ws, r, 4, pct, fmt=fmt_pct, color=cor, bg=bg, align="right")
            r += 1

        r += 1
        if inc_det and movs:
            ws2 = wb.create_sheet("Movimentações")
            ws2.sheet_view.showGridLines = False
            for ci, w in zip(range(1, 8), [3,12,12,30,34,18,3]):
                ws2.column_dimensions[get_column_letter(ci)].width = w
            titulo_sheet(ws2, f"Movimentações — {mes_nome}/{ano_ini}",
                         f"Total: {len(movs)} registros  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=5)
            tabela_movimentacoes(ws2, movs, 5, f"{mes_nome}/{ano_ini}")

        rodape(ws, r)
        nome_arquivo = f"Santos_Assoc_Mensal_{mes_nome}_{ano_ini}.xlsx"

    # ════════════════════════════════════════════════════════════
    #  TIPO 2 — COMPARATIVO DE PERÍODO (mês a mês em colunas)
    # ════════════════════════════════════════════════════════════
    elif tipo_rel == "periodo":
        # Montar lista de meses no intervalo
        periodos = []
        am, aa = mes_ini, ano_ini
        while (aa, am) <= (ano_fim, mes_fim):
            periodos.append((am, aa))
            am += 1
            if am > 12: am = 1; aa += 1

        ncols_dados = len(periodos)
        ncols_total = ncols_dados + 2  # B=categoria, depois meses, depois total

        ws = wb.active; ws.title = "Comparativo"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 32
        for ci in range(3, 3 + ncols_dados + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16

        periodo_label = f"{MESES_PT[mes_ini][:3]}/{ano_ini} → {MESES_PT[mes_fim][:3]}/{ano_fim}"
        titulo_sheet(ws, "Comparativo de Período", f"{periodo_label}  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=ncols_total)

        # Cabeçalhos
        r = 5
        ws.row_dimensions[r].height = 22
        header_cell(ws, r, 2, "CATEGORIA", bg=AZ, size=9)
        for ci, (m, a) in enumerate(periodos, 3):
            header_cell(ws, r, ci, f"{MESES_PT[m][:3]}/{a}", bg=AZ, size=9)
        header_cell(ws, r, 3 + ncols_dados, "TOTAL", bg=DOC, size=9)

        # Linhas de receitas
        r += 1
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:{get_column_letter(3+ncols_dados)}{r}")
        header_cell(ws, r, 2, "▲  RECEITAS", bg=VD, size=9)
        r += 1

        cats_e = [c["categoria"] for c in
                  _get_cats_periodo(mes_ini, ano_ini, mes_fim, ano_fim, "Entrada")]
        for cat in cats_e:
            ws.row_dimensions[r].height = 18
            data_cell(ws, r, 2, cat, bg="F0FFF4")
            vals_col = []
            for ci, (m, a) in enumerate(periodos, 3):
                v = _soma_cat(cat, "Entrada", m, a, m, a)
                data_cell(ws, r, ci, v, fmt=fmt_brl, color=VD, bg="F0FFF4", align="right")
                vals_col.append(v)
            data_cell(ws, r, 3+ncols_dados, sum(vals_col), fmt=fmt_brl,
                      color=VD, bold=True, bg="D1FAE5", align="right")
            r += 1

        # Total receitas por mês
        ws.row_dimensions[r].height = 22
        data_cell(ws, r, 2, "TOTAL RECEITAS", bold=True, bg="D1FAE5")
        tots_e = []
        for ci, (m, a) in enumerate(periodos, 3):
            v = soma_periodo("Entrada", m, a, m, a)
            tots_e.append(v)
            data_cell(ws, r, ci, v, fmt=fmt_brl, bold=True, color=VD, bg="D1FAE5", align="right")
        data_cell(ws, r, 3+ncols_dados, sum(tots_e), fmt=fmt_brl, bold=True, color=VD, bg="D1FAE5", align="right")
        r += 2

        # Linhas de despesas
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:{get_column_letter(3+ncols_dados)}{r}")
        header_cell(ws, r, 2, "▼  DESPESAS", bg=VM, size=9)
        r += 1

        cats_s = [c["categoria"] for c in
                  _get_cats_periodo(mes_ini, ano_ini, mes_fim, ano_fim, "Saida")]
        for cat in cats_s:
            ws.row_dimensions[r].height = 18
            data_cell(ws, r, 2, cat, bg="FFF5F5")
            vals_col = []
            for ci, (m, a) in enumerate(periodos, 3):
                v = _soma_cat(cat, "Saida", m, a, m, a)
                data_cell(ws, r, ci, v, fmt=fmt_brl, color=VM, bg="FFF5F5", align="right")
                vals_col.append(v)
            data_cell(ws, r, 3+ncols_dados, sum(vals_col), fmt=fmt_brl,
                      color=VM, bold=True, bg="FEE2E2", align="right")
            r += 1

        # Total despesas por mês
        ws.row_dimensions[r].height = 22
        data_cell(ws, r, 2, "TOTAL DESPESAS", bold=True, bg="FEE2E2")
        tots_s = []
        for ci, (m, a) in enumerate(periodos, 3):
            v = soma_periodo("Saida", m, a, m, a)
            tots_s.append(v)
            data_cell(ws, r, ci, v, fmt=fmt_brl, bold=True, color=VM, bg="FEE2E2", align="right")
        data_cell(ws, r, 3+ncols_dados, sum(tots_s), fmt=fmt_brl, bold=True, color=VM, bg="FEE2E2", align="right")
        r += 2

        # Lucro líquido por mês
        ws.row_dimensions[r].height = 24
        data_cell(ws, r, 2, "LUCRO LÍQUIDO", bold=True, bg="DBEAFE")
        for ci, (te, ts) in enumerate(zip(tots_e, tots_s), 3):
            liq = te - ts
            cor = "1E40AF" if liq >= 0 else VM
            data_cell(ws, r, ci, liq, fmt=fmt_brl, bold=True, color=cor, bg="DBEAFE", align="right")
        liq_total = sum(tots_e) - sum(tots_s)
        cor_lt = "1E40AF" if liq_total >= 0 else VM
        data_cell(ws, r, 3+ncols_dados, liq_total, fmt=fmt_brl, bold=True, color=cor_lt, bg="DBEAFE", align="right")
        r += 1

        # Margem
        ws.row_dimensions[r].height = 20
        data_cell(ws, r, 2, "MARGEM LÍQUIDA (%)", bold=False, bg="EFF6FF")
        for ci, (te, ts) in enumerate(zip(tots_e, tots_s), 3):
            marg = ((te-ts)/te*100) if te > 0 else 0
            cor = "1E40AF" if marg >= 0 else VM
            data_cell(ws, r, ci, marg, fmt=fmt_pct, bold=False, color=cor, bg="EFF6FF", align="right")
        marg_tot = (liq_total/sum(tots_e)*100) if sum(tots_e) > 0 else 0
        cor_mt = "1E40AF" if marg_tot >= 0 else VM
        data_cell(ws, r, 3+ncols_dados, marg_tot, fmt=fmt_pct, bold=True, color=cor_mt, bg="EFF6FF", align="right")
        r += 2

        rodape(ws, r, ncols=ncols_total)

        if inc_det:
            movs = query_movs(mes_ini, ano_ini, mes_fim, ano_fim,
                              filtro_tipo or None, cats_sel or None, ordenar)
            if movs:
                ws2 = wb.create_sheet("Movimentações")
                ws2.sheet_view.showGridLines = False
                for ci, w in zip(range(1, 8), [3,12,12,30,34,18,3]):
                    ws2.column_dimensions[get_column_letter(ci)].width = w
                titulo_sheet(ws2, f"Movimentações — {periodo_label}",
                             f"{len(movs)} registros  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=5)
                tabela_movimentacoes(ws2, movs, 5, periodo_label)

        nome_arquivo = f"Santos_Assoc_Comparativo_{MESES_PT[mes_ini][:3]}{ano_ini}_{MESES_PT[mes_fim][:3]}{ano_fim}.xlsx"

    # ════════════════════════════════════════════════════════════
    #  TIPO 3 — RELATÓRIO ANUAL (resumo + meses em abas separadas)
    # ════════════════════════════════════════════════════════════
    elif tipo_rel == "anual":
        ano = ano_ini
        ws = wb.active; ws.title = "Resumo Anual"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 30
        for ci in range(3, 16):
            ws.column_dimensions[get_column_letter(ci)].width = 13

        titulo_sheet(ws, f"Relatório Anual {ano}",
                     f"Janeiro a Dezembro de {ano}  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=14)

        # Cabeçalhos: categoria + 12 meses + total
        r = 5
        ws.row_dimensions[r].height = 22
        header_cell(ws, r, 2, "CATEGORIA", bg=AZ, size=9)
        for m in range(1, 13):
            header_cell(ws, r, 2+m, MESES_PT[m][:3], bg=AZ, size=9)
        header_cell(ws, r, 15, "TOTAL", bg=DOC, size=9)
        r += 1

        def linha_anual(ws, r, label, tipo_cat, ano, bg_row, cor_val, bg_tot):
            ws.row_dimensions[r].height = 18
            data_cell(ws, r, 2, label, bg=bg_row)
            vals = []
            for m in range(1, 13):
                v = _soma_cat(label, tipo_cat, m, ano, m, ano)
                data_cell(ws, r, 2+m, v, fmt=fmt_brl, color=cor_val, bg=bg_row, align="right")
                vals.append(v)
            data_cell(ws, r, 15, sum(vals), fmt=fmt_brl, bold=True, color=cor_val, bg=bg_tot, align="right")
            return r + 1, vals

        # Receitas
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:O{r}")
        header_cell(ws, r, 2, "▲  RECEITAS", bg=VD, size=9); r += 1
        cats_e = [c["categoria"] for c in _get_cats_periodo(1, ano, 12, ano, "Entrada")]
        tot_e_meses = [0]*12
        for cat in cats_e:
            r, vals = linha_anual(ws, r, cat, "Entrada", ano, "F0FFF4", VD, "D1FAE5")
            tot_e_meses = [tot_e_meses[i]+vals[i] for i in range(12)]
        ws.row_dimensions[r].height = 22
        data_cell(ws, r, 2, "TOTAL RECEITAS", bold=True, bg="D1FAE5")
        for m in range(12):
            data_cell(ws, r, 3+m, tot_e_meses[m], fmt=fmt_brl, bold=True, color=VD, bg="D1FAE5", align="right")
        data_cell(ws, r, 15, sum(tot_e_meses), fmt=fmt_brl, bold=True, color=VD, bg="D1FAE5", align="right")
        r += 2

        # Despesas
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:O{r}")
        header_cell(ws, r, 2, "▼  DESPESAS", bg=VM, size=9); r += 1
        cats_s = [c["categoria"] for c in _get_cats_periodo(1, ano, 12, ano, "Saida")]
        tot_s_meses = [0]*12
        for cat in cats_s:
            r, vals = linha_anual(ws, r, cat, "Saida", ano, "FFF5F5", VM, "FEE2E2")
            tot_s_meses = [tot_s_meses[i]+vals[i] for i in range(12)]
        ws.row_dimensions[r].height = 22
        data_cell(ws, r, 2, "TOTAL DESPESAS", bold=True, bg="FEE2E2")
        for m in range(12):
            data_cell(ws, r, 3+m, tot_s_meses[m], fmt=fmt_brl, bold=True, color=VM, bg="FEE2E2", align="right")
        data_cell(ws, r, 15, sum(tot_s_meses), fmt=fmt_brl, bold=True, color=VM, bg="FEE2E2", align="right")
        r += 2

        # Lucro por mês
        ws.row_dimensions[r].height = 24
        data_cell(ws, r, 2, "LUCRO LÍQUIDO", bold=True, bg="DBEAFE")
        for m in range(12):
            liq = tot_e_meses[m] - tot_s_meses[m]
            cor = "1E40AF" if liq >= 0 else VM
            data_cell(ws, r, 3+m, liq, fmt=fmt_brl, bold=True, color=cor, bg="DBEAFE", align="right")
        liq_ano = sum(tot_e_meses) - sum(tot_s_meses)
        data_cell(ws, r, 15, liq_ano, fmt=fmt_brl, bold=True,
                  color="1E40AF" if liq_ano >= 0 else VM, bg="DBEAFE", align="right")
        r += 1

        # Margem
        ws.row_dimensions[r].height = 20
        data_cell(ws, r, 2, "MARGEM LÍQUIDA (%)", bg="EFF6FF")
        for m in range(12):
            marg = ((tot_e_meses[m]-tot_s_meses[m])/tot_e_meses[m]*100) if tot_e_meses[m] > 0 else 0
            data_cell(ws, r, 3+m, marg, fmt=fmt_pct, color="1E40AF" if marg >= 0 else VM, bg="EFF6FF", align="right")
        marg_ano = (liq_ano/sum(tot_e_meses)*100) if sum(tot_e_meses) > 0 else 0
        data_cell(ws, r, 15, marg_ano, fmt=fmt_pct, bold=True,
                  color="1E40AF" if marg_ano >= 0 else VM, bg="EFF6FF", align="right")
        r += 2
        rodape(ws, r, ncols=14)

        if inc_det:
            movs = query_movs(1, ano, 12, ano, filtro_tipo or None, cats_sel or None, ordenar)
            if movs:
                ws2 = wb.create_sheet("Todas as Movimentações")
                ws2.sheet_view.showGridLines = False
                for ci, w in zip(range(1, 8), [3,12,12,30,34,18,3]):
                    ws2.column_dimensions[get_column_letter(ci)].width = w
                titulo_sheet(ws2, f"Movimentações — {ano}",
                             f"{len(movs)} registros  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=5)
                tabela_movimentacoes(ws2, movs, 5, str(ano))

        nome_arquivo = f"Santos_Assoc_Anual_{ano}.xlsx"

    # ════════════════════════════════════════════════════════════
    #  TIPO 4 — EXTRATO DE FLUXO DE CAIXA (cronológico)
    # ════════════════════════════════════════════════════════════
    elif tipo_rel == "fluxo":
        periodo_label = f"{MESES_PT[mes_ini][:3]}/{ano_ini} → {MESES_PT[mes_fim][:3]}/{ano_fim}"
        movs = query_movs(mes_ini, ano_ini, mes_fim, ano_fim,
                          filtro_tipo or None, cats_sel or None, "data ASC, id ASC")

        ws = wb.active; ws.title = "Fluxo de Caixa"
        ws.sheet_view.showGridLines = False
        for ci, w in zip(range(1, 9), [3,12,12,30,30,16,16,3]):
            ws.column_dimensions[get_column_letter(ci)].width = w

        titulo_sheet(ws, "Extrato — Fluxo de Caixa",
                     f"{periodo_label}  •  Gerado em {hoje.strftime('%d/%m/%Y')}", ncols=6)

        r = 5
        ws.row_dimensions[r].height = 22
        for ci, h in zip(range(2, 8), ["Data","Tipo","Categoria","Descrição","Valor (R$)","Saldo Acum. (R$)"]):
            header_cell(ws, r, ci, h, size=9)
        r += 1

        saldo = 0.0
        for m in movs:
            ws.row_dimensions[r].height = 18
            saldo += m["valor"] if m["tipo"] == "Entrada" else -m["valor"]
            bg = "F0FFF4" if m["tipo"] == "Entrada" else "FFF5F5"
            cor = VD if m["tipo"] == "Entrada" else VM
            data_cell(ws, r, 2, datetime.strptime(m["data"],"%Y-%m-%d").strftime("%d/%m/%Y"), bg=bg, align="center")
            data_cell(ws, r, 3, m["tipo"], bold=True, color=cor, bg=bg, align="center")
            data_cell(ws, r, 4, m["categoria"], bg=bg)
            data_cell(ws, r, 5, m["descricao"], bg=bg)
            data_cell(ws, r, 6, m["valor"] if m["tipo"] == "Entrada" else -m["valor"],
                      fmt=fmt_brl, bold=True, color=cor, bg=bg, align="right")
            cor_s = "1E40AF" if saldo >= 0 else VM
            data_cell(ws, r, 7, saldo, fmt=fmt_brl, bold=True, color=cor_s, bg="F0F4FF", align="right")
            r += 1

        r += 1
        rodape(ws, r, ncols=6)
        nome_arquivo = f"Santos_Assoc_FluxoCaixa_{MESES_PT[mes_ini][:3]}{ano_ini}_{MESES_PT[mes_fim][:3]}{ano_fim}.xlsx"

    else:
        nome_arquivo = "Santos_Assoc_Relatorio.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Helpers internos de consulta por categoria
def _get_cats_periodo(m_ini, a_ini, m_fim, a_fim, tipo):
    with get_db() as conn:
        return conn.execute(
            "SELECT DISTINCT categoria FROM movimentacoes "
            "WHERE tipo=? AND data >= ? AND data <= ? ORDER BY categoria",
            (tipo, f"{a_ini}-{m_ini:02d}-01", f"{a_fim}-{m_fim:02d}-31")
        ).fetchall()

def _soma_cat(cat, tipo, m_ini, a_ini, m_fim, a_fim):
    with get_db() as conn:
        r = conn.execute(
            "SELECT COALESCE(SUM(valor),0) AS t FROM movimentacoes "
            "WHERE categoria=? AND tipo=? AND data >= ? AND data <= ?",
            (cat, tipo, f"{a_ini}-{m_ini:02d}-01", f"{a_fim}-{m_fim:02d}-31")
        ).fetchone()
    return r["t"]


# ----------------------------------------------------------
# INICIALIZAÇÃO
# ----------------------------------------------------------

if __name__ == "__main__":
    init_db()
    print("\n✅  Santos & Associados – Sistema Financeiro")
    print("   Acesse: http://127.0.0.1:5000\n")
    app.run(debug=True)
